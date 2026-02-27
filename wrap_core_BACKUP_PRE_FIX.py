from __future__ import annotations

import os
import tempfile
from datetime import datetime

import streamlit as st
import pandas as pd

from logistat_engine import run_logistat_v11, RULES
from config_store import load_config, save_config

# Branding (opcional)
try:
    from branding import COMPANY_NAME, LOGISTAT_VERSION, AUTHOR_NAME, AUTHOR_ROLE
except Exception:
    COMPANY_NAME = ""
    LOGISTAT_VERSION = ""
    AUTHOR_NAME = ""
    AUTHOR_ROLE = ""

def _save_uploaded_to_tmp(uploaded_file, suffix: str) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getbuffer())
        return tmp.name

def _peek_plan_df(plan_file) -> pd.DataFrame | None:
    try:
        xls = pd.ExcelFile(plan_file)
        sheet = "PLAN" if "PLAN" in xls.sheet_names else xls.sheet_names[0]
        return pd.read_excel(xls, sheet_name=sheet)
    except Exception:
        return None

def _postprocess_excel_info(excel_path: str) -> None:
    """Agrega hoja INFO institucional si existe openpyxl."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        if "INFO" in wb.sheetnames:
            ws = wb["INFO"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("INFO", 0)

        rows = [
            ("Empresa", COMPANY_NAME),
            ("Versión", LOGISTAT_VERSION),
            ("Elaboró", AUTHOR_NAME),
            ("Cargo", AUTHOR_ROLE),
            ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for i, (k, v) in enumerate(rows, start=1):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 90
        wb.save(excel_path)
    except Exception:
        return

def render():
    st.header("🚚 Planeación")

    base_dir = os.getcwd()
    cfg = load_config(base_dir)

    st.caption("Sube el PLAN (Excel) y opcionalmente el template de Diésel. Ajusta parámetros y ejecuta.")

    col1, col2 = st.columns([0.6, 0.4])

    with col1:
        plan_file = st.file_uploader(
            "PLAN (.xlsx) €” columnas mínimas: cliente, obra, hora_solicitada, lat, lon, localidad, m3 (bombeable opcional)",
            type=["xlsx"],
            key="plan_up",
        )
        diesel_file = st.file_uploader(
            "Diésel (opcional) €” template con hojas dim_revolvedora y fact_revolvedoras",
            type=["xlsx"],
            key="diesel_up",
        )

    plan_df = _peek_plan_df(plan_file) if plan_file is not None else None

    with col2:
        st.subheader("Parámetros")
        st.caption("Se guardan en config.json")

        planta_lat = st.number_input("PLANTA_LAT", value=float(cfg.get("PLANTA_LAT", RULES.get("PLANTA_LAT", 0.0))), format="%.6f")
        planta_lon = st.number_input("PLANTA_LON", value=float(cfg.get("PLANTA_LON", RULES.get("PLANTA_LON", 0.0))), format="%.6f")
        cap_turno = st.number_input("PLANTA_CAP_M3_TURNO", value=float(cfg.get("PLANTA_CAP_M3_TURNO", RULES.get("PLANTA_CAP_M3_TURNO", 230.0))))
        n_ollas = st.number_input("N_OLLAS_DISPONIBLES", value=int(cfg.get("N_OLLAS_DISPONIBLES", RULES.get("N_OLLAS_DISPONIBLES", 11))), step=1, min_value=1)
        cap_olla = st.number_input("CAPACIDAD_OLLA_M3", value=float(cfg.get("CAPACIDAD_OLLA_M3", RULES.get("CAPACIDAD_OLLA_M3", 7.0))))

        cA, cB = st.columns(2)
        with cA:
            if st.button("💾 Guardar parámetros", use_container_width=True):
                save_config(base_dir, {
                    "PLANTA_LAT": float(planta_lat),
                    "PLANTA_LON": float(planta_lon),
                    "PLANTA_CAP_M3_TURNO": float(cap_turno),
                    "N_OLLAS_DISPONIBLES": int(n_ollas),
                    "CAPACIDAD_OLLA_M3": float(cap_olla),
                })
                st.success("Guardado en config.json")

        with cB:
            if st.button("🧭 Autodetectar planta (promedio obras)", use_container_width=True, disabled=(plan_df is None)):
                if plan_df is not None:
                    cols = {c.lower(): c for c in plan_df.columns}
                    if "lat" in cols and "lon" in cols:
                        lat_mean = float(pd.to_numeric(plan_df[cols["lat"]], errors="coerce").dropna().mean())
                        lon_mean = float(pd.to_numeric(plan_df[cols["lon"]], errors="coerce").dropna().mean())
                        st.session_state["__auto_planta_lat"] = lat_mean
                        st.session_state["__auto_planta_lon"] = lon_mean
                        st.warning("Autodetección aplicada temporalmente. Revisa y luego guarda si te sirve.")
                    else:
                        st.error("No encontré columnas lat/lon en el PLAN.")

        if "__auto_planta_lat" in st.session_state:
            planta_lat = float(st.session_state["__auto_planta_lat"])
            planta_lon = float(st.session_state["__auto_planta_lon"])
            st.info(f"Temporal: PLANTA_LAT/LON = {planta_lat:.6f}, {planta_lon:.6f}")

    if float(planta_lat) == 0.0 and float(planta_lon) == 0.0:
        st.warning("PLANTA_LAT/LON están en 0,0. Ajusta antes de ejecutar (o usa autodetectar).")

    st.divider()

    if plan_df is not None:
        required = ["cliente", "obra", "hora_solicitada", "lat", "lon", "localidad", "m3"]
        cols_lower = {c.lower(): c for c in plan_df.columns}
        missing = [r for r in required if r not in cols_lower]
        if missing:
            st.error(f"Faltan columnas obligatorias en PLAN: {', '.join(missing)}")
        else:
            with st.expander("Vista rápida PLAN"):
                st.dataframe(plan_df.head(25), use_container_width=True)

    disabled_run = (plan_file is None) or (float(planta_lat) == 0.0 and float(planta_lon) == 0.0)

    if st.button("–¶ï¸ Ejecutar LOGISTAT", use_container_width=True, disabled=disabled_run):
        try:
            plan_path = _save_uploaded_to_tmp(plan_file, ".xlsx")
            diesel_path = _save_uploaded_to_tmp(diesel_file, ".xlsx") if diesel_file is not None else None

            overrides = {
                "PLANTA_LAT": float(planta_lat),
                "PLANTA_LON": float(planta_lon),
                "PLANTA_CAP_M3_TURNO": float(cap_turno),
                "N_OLLAS_DISPONIBLES": int(n_ollas),
                "CAPACIDAD_OLLA_M3": float(cap_olla),
            }

            result = run_logistat_v11(
                plan_path=plan_path,
                diesel_template_path=diesel_path,
                base_dir=base_dir,
                rules_overrides=overrides,
            )

            st.session_state["result"] = result
            st.session_state["out_xlsx"] = result.get("plan_excel")

            # INFO institucional en Excel principal
            if isinstance(st.session_state.get("out_xlsx"), str) and os.path.exists(st.session_state["out_xlsx"]):
                _postprocess_excel_info(st.session_state["out_xlsx"])

            # Exportaciones: Global, Producción (PlanV9), Diésel
            from exports_global import run_global_exports
            from exports_planv9 import run_planv9_exports
            from exports_diesel import run_diesel_exports

            plan_date = result.get("plan_date")
            run_global_exports(base_dir, st.session_state.get("out_xlsx"))
            run_planv9_exports(base_dir, result, st.session_state.get("out_xlsx"), plan_date=plan_date)
            run_diesel_exports(base_dir, result, st.session_state.get("out_xlsx"), plan_date=plan_date, diesel_uploaded=(diesel_file is not None))

            st.success("Listo. Revisa Dashboard / Exportación / Reportes.")

            with st.expander("Resumen"):
                st.dataframe(result.get("resumen"), use_container_width=True)
            with st.expander("Alertas"):
                st.dataframe(result.get("alertas_plan"), use_container_width=True)
            with st.expander("Plan calculado (primeras filas)"):
                pc = result.get("plan_calculado")
                st.dataframe(pc.head(200) if hasattr(pc, "head") else pc, use_container_width=True)

        except Exception as e:
            st.error(f"Error al ejecutar: {e}")

    if st.button("🧹 Limpiar sesión", use_container_width=True):
        for k in ["result", "out_xlsx", "out_global_full", "out_global_sheets",
                  "out_planv9_full", "out_planv9_sheets",
                  "out_diesel_full", "out_diesel_sheets",
                  "__auto_planta_lat", "__auto_planta_lon"]:
            if k in st.session_state:
                del st.session_state[k]
        st.success("Sesión limpia.")




